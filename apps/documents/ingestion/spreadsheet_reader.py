import csv
from io import BytesIO, StringIO
from typing import Dict, List, Tuple

from openpyxl import load_workbook

SPREADSHEET_SCHEMA_VERSION = "v1"


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_text_from_map(cell_map: Dict[str, str]) -> str:
    pairs = [f"{k}={v}" for k, v in cell_map.items() if v]
    if pairs:
        return " ; ".join(pairs)
    return ""


def _sheet_to_canonical(name: str, rows: List[List[str]]) -> Dict:
    if not rows:
        return {"name": name, "columns": [], "rows": []}

    header = [h.strip() for h in rows[0]]
    data_rows = rows[1:] if any(header) else rows
    row_start = 2 if any(header) else 1

    canonical_rows: List[Dict] = []
    for offset, row in enumerate(data_rows, start=row_start):
        col_count = max(len(row), len(header))
        cells = [row[idx].strip() if idx < len(row) else "" for idx in range(col_count)]

        cell_map: Dict[str, str] = {}
        for idx, val in enumerate(cells, start=1):
            key = header[idx - 1] if idx - 1 < len(header) and header[idx - 1] else f"col_{idx}"
            cell_map[key] = val

        row_text = _row_text_from_map(cell_map)
        canonical_rows.append(
            {
                "row_number": offset,
                "cells": cells,
                "cell_map": cell_map,
                "text": row_text,
            }
        )

    return {"name": name, "columns": header if any(header) else [], "rows": canonical_rows}


def _canonical_to_text(metadata: Dict) -> str:
    parts: List[str] = []
    for sheet in metadata.get("sheets", []):
        sheet_name = sheet.get("name") or "Sheet"
        parts.append(f"[Sheet: {sheet_name}]")
        for row in sheet.get("rows", []):
            row_num = row.get("row_number")
            row_text = row.get("text") or ""
            if row_text:
                parts.append(f"Row {row_num}: {row_text}")
        parts.append("")
    return "\n".join(parts).strip()


def parse_csv_bytes(raw: bytes) -> Tuple[str, Dict]:
    decoded = raw.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(StringIO(decoded))
    rows = [[_normalize_cell(cell) for cell in row] for row in reader]

    metadata = {
        "kind": "spreadsheet",
        "schema_version": SPREADSHEET_SCHEMA_VERSION,
        "sheets": [_sheet_to_canonical("Sheet1", rows)],
    }
    return _canonical_to_text(metadata), metadata


def parse_xlsx_bytes(raw: bytes) -> Tuple[str, Dict]:
    wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    sheets: List[Dict] = []
    for ws in wb.worksheets:
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_normalize_cell(cell) for cell in row])
        sheets.append(_sheet_to_canonical(ws.title, rows))

    metadata = {
        "kind": "spreadsheet",
        "schema_version": SPREADSHEET_SCHEMA_VERSION,
        "sheets": sheets,
    }
    return _canonical_to_text(metadata), metadata
